import os
from openpyxl import load_workbook
import openpyxl as openpyxl
import requests
# current script's directory
script_dir = os.path.dirname(os.path.realpath(__file__))
# absolute path to file
excel_file_path = os.path.join(script_dir, 'data.xlsx')
work_book = openpyxl.load_workbook(excel_file_path)


def get_create_user_data_from_excel():
    sheet = work_book['user_data']
    user_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        user_dict = {
            'name': row[0],
            'gender': row[1],
            'email': row[2],
            'status': row[3],
        }
        user_data.append(user_dict)
    return user_data

